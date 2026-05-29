"""
테라로사 외부몰 — 세트 상품 관리 페이지
pages/세트상품관리.py 로 저장하세요.
"""

import json
import re
import io
from copy import copy
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook

# GitHub 영구 저장소
try:
    from github_storage import gh_load, gh_save
    _USE_GITHUB = True
except Exception:
    _USE_GITHUB = False
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 상수 (기존 app.py와 동일)
# ──────────────────────────────────────────────
DRIP_GREEN   = "E2EFDA"
SET_BLUE     = "DDEEFF"
HEADER_GRAY  = "D9D9D9"
BORDER_COLOR = "AAAAAA"
HEADERS      = ["품목명", "중량", "옵션", "수량", "상품코드"]
COL_WIDTHS   = [42, 10, 35, 8, 18]
GROUP_ORDER  = {"세트": 0, "기타": 1, "드립백": 2, "원두": 3}
WEIGHT_ORDER = {"1kg": 0, "500g": 1, "250g": 2, "150g": 3}
WEIGHT_PAT   = re.compile(r"^(\d+(?:\.\d+)?(?:kg|g))$", re.IGNORECASE)

OPTION_REMOVE = [
    "/구매 안함", "/플러스", "[플러스] ", "/불필요", "/필요",
    "불필요", "필요", "/상자 없음", ":상자 없음", "상자 없음",
    "/포장 없음", "테라로사 시그니처 ", "[Online Exclusive] ",
    "[Online Exclusive/플러스] ",
]
OPTION_REPLACE = {
    "중간 분쇄(드립용)": "드립용",
    "드립&커피메이커":   "드립용",
    "가는 분쇄(에스프레소용)": "에스프레소용",
}

CONFIG_PATH = Path("set_config_sabang.json")

# ──────────────────────────────────────────────
# 세트 설정 로드/저장
# ──────────────────────────────────────────────
def load_set_config() -> dict:
    if _USE_GITHUB:
        return gh_load("set_config_sabang.json", {})
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_set_config(cfg: dict):
    if _USE_GITHUB:
        gh_save("set_config_sabang.json", cfg)
    else:
        CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

# ──────────────────────────────────────────────
# 세트 분리
# ──────────────────────────────────────────────
def expand_set_items(df: pd.DataFrame, set_config: dict) -> pd.DataFrame:
    if not set_config:
        return df
    expanded = []
    for _, row in df.iterrows():
        name = str(row.get("품목명", "")).strip()
        matched = None
        for set_name, components in set_config.items():
            if set_name in name:
                matched = (set_name, components)
                break
        if matched and matched[1]:
            set_name, components = matched
            qty = float(row.get("수량", 1))
            for comp in components:
                new_row = row.copy()
                new_row["품목명"] = comp["name"]
                new_row["중량"]   = comp.get("weight", "")
                new_row["옵션"]   = comp.get("option", "")
                new_row["수량"]   = qty * comp.get("qty", 1)
                new_row["_is_set"] = True
                expanded.append(new_row)
        else:
            row = row.copy()
            if "_is_set" not in row.index:
                row["_is_set"] = False
            expanded.append(row)
    return pd.DataFrame(expanded).reset_index(drop=True)

# ──────────────────────────────────────────────
# 기존 처리 로직 (app.py 그대로)
# ──────────────────────────────────────────────
def _border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

def _header(cell):
    cell.font      = Font(name="Arial", size=10, bold=True)
    cell.fill      = PatternFill("solid", start_color=HEADER_GRAY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _border()

def _data(cell, bg=None, center=False):
    cell.font      = Font(name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg) if bg else PatternFill()
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border    = _border()

def _blank(ws, row, ncols=5):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c, value="").border = _border()

def _clean_name(name: str) -> str:
    name = re.sub(r"\[테라로사\]\s*", "", str(name).strip())
    name = re.sub(r"\s+\d+(?:\.\d+)?(?:kg|g)$", "", name.strip())
    return name.strip()

def _clean_option(opt: str) -> str:
    opt = str(opt).strip()
    for r in OPTION_REMOVE:
        opt = opt.replace(r, "")
    for old, new in OPTION_REPLACE.items():
        opt = opt.replace(old, new)
    opt = opt.strip()
    return "" if opt == "단품" else opt

def _split_weight(raw_name: str, option: str):
    weight, opt_out = "", option
    if ":" in option:
        parts = [p.strip() for p in option.split(":")]
        wp, np_ = [], []
        for p in parts:
            (wp if WEIGHT_PAT.match(p) else np_).append(p)
        if wp:
            weight  = wp[0]
            opt_out = ":".join(np_)
    else:
        m = re.search(r"(\d+(?:\.\d+)?(?:kg|g))", option, re.IGNORECASE)
        if m:
            weight  = m.group(1)
            opt_out = option.replace(weight, "").strip()
    if not weight:
        m2 = re.search(r"\s+(\d+(?:\.\d+)?(?:kg|g))$", str(raw_name).strip(), re.IGNORECASE)
        if m2:
            weight = m2.group(1)
    return weight, opt_out.strip().strip(":").strip()

def _classify(name: str, weight: str) -> str:
    if "드립백" in name: return "드립백"
    if "세트"  in name: return "세트"
    if weight and re.search(r"\d+(?:kg|g)", str(weight), re.IGNORECASE): return "원두"
    return "기타"

def _sort(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_g"] = df["분류"].map(GROUP_ORDER)
    df["_w"] = df["중량"].apply(lambda w: WEIGHT_ORDER.get(str(w), 99))
    df = df.sort_values(["_g", "품목명", "_w", "옵션"])
    return df.drop(columns=["_g", "_w"]).reset_index(drop=True)

def _build_matcher(codes_df: pd.DataFrame):
    rows = []
    for _, r in codes_df.iterrows():
        cp = re.sub(r"\s+", " ", re.sub(r"\[테라로사\]\s*", "", str(r["상품명"]))).strip()
        rows.append((cp, str(r["옵션"]).strip(), str(r["코드"])))

    def match(product, weight, option):
        p = re.sub(r"\s+", " ", re.sub(r"\[테라로사\]\s*", "", str(product))).strip()
        w, o = str(weight).strip(), str(option).strip()
        results = []
        for cp, cw, code in rows:
            exact   = cp == p
            partial = (p in cp) or (cp in p)
            wm, om  = (cw == w if w else False), (cw == o if o else False)
            if exact:
                if w and wm:   results.append((1, code))
                elif o and om: results.append((2, code))
                else:          results.append((3, code))
            elif partial and (wm or om):
                results.append((4, code))
        return min(results)[1] if results else ""

    return match

def _agg(df: pd.DataFrame, include_channel: bool) -> pd.DataFrame:
    def key(r):
        no_weight = r["분류"] in ("세트", "드립백")
        parts = ([r["채널"]] if include_channel else []) + \
                [r["분류"], r["품목명"], "" if no_weight else r["중량"], r["옵션"]]
        return tuple(parts)
    df = df.copy()
    df["_k"] = df.apply(key, axis=1)
    agg_cols = dict(채널=("채널","first"), 품목명=("품목명","first"),
                    중량=("중량","first"), 옵션=("옵션","first"),
                    수량=("수량","sum"),   분류=("분류","first"))
    if "_is_set" in df.columns:
        agg_cols["_is_set"] = ("_is_set", "first")
    return df.groupby("_k", sort=False).agg(**agg_cols).reset_index(drop=True)

def _sheet1(ws, df):
    ws.title = "전체 주문취합"
    for c, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        _header(ws.cell(row=1, column=c, value=h))
        ws.column_dimensions[get_column_letter(c)].width = w
    row, prev = 2, None
    for _, r in df.iterrows():
        if prev is not None and r["품목명"] != prev:
            _blank(ws, row); row += 1
        is_set = r.get("_is_set", False)
        bg = SET_BLUE if is_set else (DRIP_GREEN if r["분류"] == "드립백" else None)
        vals = [r["품목명"], r["중량"], r["옵션"], int(r["수량"]), r["상품코드"]]
        for c, v in enumerate(vals, 1):
            _data(ws.cell(row=row, column=c, value=v or ""), bg=bg, center=(c in [2,4,5]))
        row += 1; prev = r["품목명"]

def _sheet2(ws, df):
    ws.title = "원두 중량 합산"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 15
    for c, h in enumerate(["품목명", "중량(kg)"], 1):
        _header(ws.cell(row=1, column=c, value=h))
    def to_kg(w, q):
        m = re.match(r"(\d+(?:\.\d+)?)(kg|g)", str(w).lower())
        if not m: return 0
        v = float(m.group(1))
        return (v if m.group(2)=="kg" else v/1000) * q
    beans = df[df["분류"]=="원두"].copy()
    beans["_kg"] = beans.apply(lambda r: to_kg(r["중량"], r["수량"]), axis=1)
    for i, (_, r) in enumerate(beans.groupby("품목명")["_kg"].sum().reset_index().sort_values("품목명").iterrows(), 2):
        _data(ws.cell(row=i, column=1, value=r["품목명"]))
        _data(ws.cell(row=i, column=2, value=round(r["_kg"],3)), center=True)

def _sheet3(ws, df, ch_order):
    ws.title = "채널별 주문취합"
    for c, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        _header(ws.cell(row=1, column=c, value=h))
        ws.column_dimensions[get_column_letter(c)].width = w
    row, first = 2, True
    for ch in ch_order:
        sub = _sort(df[df["채널"]==ch].copy())
        if sub.empty: continue
        if not first: _blank(ws, row); row += 1
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c, value=ch if c==1 else "")
            cell.font      = Font(name="Arial", size=10, bold=True)
            cell.fill      = PatternFill("solid", start_color=HEADER_GRAY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = _border()
        row += 1; prev = None
        for _, r in sub.iterrows():
            if prev is not None and r["품목명"] != prev:
                _blank(ws, row); row += 1
            is_set = r.get("_is_set", False)
            bg = SET_BLUE if is_set else (DRIP_GREEN if r["분류"]=="드립백" else None)
            vals = [r["품목명"], r["중량"], r["옵션"], int(r["수량"]), r["상품코드"]]
            for c, v in enumerate(vals, 1):
                _data(ws.cell(row=row, column=c, value=v or ""), bg=bg, center=(c in [2,4,5]))
            row += 1; prev = r["품목명"]
        first = False

def process(orders_bytes: bytes, codes_bytes: bytes, set_config: dict) -> bytes:
    orders_df = pd.read_excel(io.BytesIO(orders_bytes), dtype=str)
    orders_df.columns = ["채널", "상품명", "옵션", "수량"]
    orders_df["수량"] = pd.to_numeric(orders_df["수량"], errors="coerce").fillna(0)

    codes_df = pd.read_excel(io.BytesIO(codes_bytes), dtype=str)
    codes_df.columns = ["코드", "상품명", "옵션"]
    codes_df = codes_df.fillna("")

    matcher  = _build_matcher(codes_df)
    ch_order = list(dict.fromkeys(orders_df["채널"].tolist()))

    rows = []
    for _, r in orders_df.iterrows():
        opt    = _clean_option(str(r["옵션"]))
        weight, opt_clean = _split_weight(str(r["상품명"]), opt)
        rows.append({"채널": r["채널"], "품목명": _clean_name(str(r["상품명"])),
                     "중량": weight, "옵션": opt_clean, "수량": r["수량"], "_is_set": False})
    df = pd.DataFrame(rows)
    df["분류"] = df.apply(lambda r: _classify(r["품목명"], r["중량"]), axis=1)

    # ★ 세트 분리 적용
    df = expand_set_items(df, set_config)
    df["분류"] = df.apply(lambda r: _classify(r["품목명"], r["중량"]), axis=1)

    total = _agg(df, include_channel=False)
    by_ch = _agg(df, include_channel=True)

    for frame in (total, by_ch):
        frame["상품코드"] = frame.apply(
            lambda r: matcher(r["품목명"], r["중량"], r["옵션"]), axis=1)

    total_s = _sort(total)
    by_ch_s = _sort(by_ch)

    wb = Workbook()
    _sheet1(wb.active, total_s)
    _sheet2(wb.create_sheet(), total_s)
    _sheet3(wb.create_sheet(), by_ch_s, ch_order)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════
st.set_page_config(page_title="세트 상품 관리", page_icon="📦", layout="wide")

st.markdown("""
<style>
[data-testid="stSidebar"] { background: #FAF3F0; }
h1, h2, h3 { color: #8B3A2A !important; }
.stButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600;
}
.stButton > button:hover { background: #C4644A; color: white; }
.stDownloadButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600; width: 100%;
}
</style>
""", unsafe_allow_html=True)

if "set_config" not in st.session_state:
    st.session_state.set_config = load_set_config()
if "editing_set" not in st.session_state:
    st.session_state.editing_set = None

st.title("📦 세트 상품 관리")
st.caption("세트 상품을 구성 품목별로 분리하여 주문취합 엑셀에 반영합니다.")
st.divider()

left, right = st.columns([1, 1], gap="large")

# ── 왼쪽: 세트 목록 ──
with left:
    st.subheader("세트 상품명")

    col_input, col_btn = st.columns([4, 1])
    with col_input:
        new_name = st.text_input("세트 상품명", placeholder="세트 상품명",
                                  label_visibility="collapsed", key="new_set_input")
    with col_btn:
        if st.button("추가", use_container_width=True, key="btn_add"):
            name = new_name.strip()
            if name and name not in st.session_state.set_config:
                st.session_state.set_config[name] = []
                save_set_config(st.session_state.set_config)
                st.session_state.editing_set = name
                st.rerun()
            elif name in st.session_state.set_config:
                st.warning("이미 등록된 세트입니다.")
            else:
                st.warning("세트 상품명을 입력하세요.")

    h1, h2, h3 = st.columns([5, 3, 2])
    with h1: st.caption("상품명")
    with h2: st.caption("등록일")
    with h3: st.caption("저장")
    st.divider()

    if not st.session_state.set_config:
        st.info("등록된 세트 상품이 없습니다.")
    else:
        for set_name in list(st.session_state.set_config.keys()):
            is_editing = st.session_state.editing_set == set_name
            c1, c2, c3, c4, c5 = st.columns([4, 3, 1, 1, 1])
            with c1: st.markdown(f"**{set_name}**")
            with c2: st.caption(datetime.today().strftime("%Y-%m-%d"))
            with c3:
                if st.button("선택" if not is_editing else "닫기",
                              key=f"sel_{set_name}", use_container_width=True):
                    st.session_state.editing_set = None if is_editing else set_name
                    st.rerun()
            with c4:
                if st.button("저장", key=f"save_{set_name}", use_container_width=True):
                    save_set_config(st.session_state.set_config)
                    st.success("저장!")
            with c5:
                if st.button("삭제", key=f"del_{set_name}", use_container_width=True):
                    del st.session_state.set_config[set_name]
                    if st.session_state.editing_set == set_name:
                        st.session_state.editing_set = None
                    save_set_config(st.session_state.set_config)
                    st.rerun()

# ── 오른쪽: 구성 옵션 편집 ──
with right:
    sel = st.session_state.editing_set
    if not sel or sel not in st.session_state.set_config:
        st.subheader("구성 옵션")
        st.info("← 왼쪽에서 세트 상품을 선택하세요.")
    else:
        st.subheader(f"구성 옵션 — {sel}")
        comps = st.session_state.set_config[sel]

        hc1, hc2, hc3, hc4, hc5, hc6, hc7 = st.columns([1, 4, 2, 2, 2, 2, 1])
        with hc1: st.caption("수량")
        with hc2: st.caption("구성 품목명")
        with hc3: st.caption("중량")
        with hc4: st.caption("옵션")
        with hc5: st.caption("등록일")
        with hc6: st.caption("저장")
        st.divider()

        for i, comp in enumerate(comps):
            rc1, rc2, rc3, rc4, rc5, rc6, rc7 = st.columns([1, 4, 2, 2, 2, 2, 1])
            with rc1:
                new_qty = st.number_input("수량", min_value=1, value=comp.get("qty", 1),
                                           key=f"qty_{sel}_{i}", label_visibility="collapsed")
            with rc2:
                new_nm = st.text_input("품목명", value=comp.get("name", ""),
                                        key=f"nm_{sel}_{i}", label_visibility="collapsed")
            with rc3:
                new_weight = st.text_input("중량", value=comp.get("weight", ""),
                                            placeholder="예: 250g",
                                            key=f"wt_{sel}_{i}", label_visibility="collapsed")
            with rc4:
                new_option = st.text_input("옵션", value=comp.get("option", ""),
                                            placeholder="옵션(선택)",
                                            key=f"op_{sel}_{i}", label_visibility="collapsed")
            with rc5:
                st.caption(datetime.today().strftime("%Y-%m-%d"))
            with rc6:
                if st.button("저장", key=f"csave_{sel}_{i}", use_container_width=True):
                    comps[i] = {"name": new_nm, "qty": int(new_qty),
                                 "weight": new_weight.strip(), "option": new_option.strip()}
                    save_set_config(st.session_state.set_config)
                    st.success("저장!")
            with rc7:
                if st.button("삭제", key=f"cdel_{sel}_{i}", use_container_width=True):
                    comps.pop(i)
                    save_set_config(st.session_state.set_config)
                    st.rerun()

        st.divider()
        na1, na2, na3, na4, na5 = st.columns([1, 4, 2, 2, 1])
        with na1:
            add_qty = st.number_input("수량", min_value=1, value=1,
                                       key=f"addqty_{sel}", label_visibility="collapsed")
        with na2:
            add_nm = st.text_input("품목명", placeholder="구성 품목명",
                                    key=f"addnm_{sel}", label_visibility="collapsed")
        with na3:
            add_weight = st.text_input("중량", placeholder="예: 250g",
                                        key=f"addwt_{sel}", label_visibility="collapsed")
        with na4:
            add_option = st.text_input("옵션", placeholder="옵션(선택)",
                                        key=f"addop_{sel}", label_visibility="collapsed")
        with na5:
            if st.button("추가", key=f"addcomp_{sel}", use_container_width=True):
                if add_nm.strip():
                    comps.append({"name": add_nm.strip(), "qty": int(add_qty),
                                   "weight": add_weight.strip(), "option": add_option.strip()})
                    save_set_config(st.session_state.set_config)
                    st.rerun()

st.divider()

# ── 주문 처리 ──
st.subheader("주문취합 처리")
col1, col2 = st.columns(2)
with col1:
    orders_file = st.file_uploader("📋 주문 파일 (.xlsx)", type=["xlsx"],
                                    help="A열:쇼핑몰명 / B열:상품명 / C열:옵션 / D열:수량")
with col2:
    codes_file  = st.file_uploader("🗂️ 상품코드 파일 (.xlsx)", type=["xlsx"])

if orders_file and codes_file:
    if st.button("🚀 취합 파일 생성", use_container_width=True):
        with st.spinner("처리 중..."):
            try:
                result_bytes = process(orders_file.read(), codes_file.read(),
                                       st.session_state.set_config)
                today = date.today().strftime("%Y%m%d")
                st.success("✅ 취합 완료!")
                if st.session_state.set_config:
                    st.info(f"📦 세트 분리 적용: {len(st.session_state.set_config)}개 → 하늘색 행으로 표시됨")
                st.download_button(
                    label="⬇️ 취합 파일 다운로드",
                    data=result_bytes,
                    file_name=f"외부몰_주문취합_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"❌ 오류: {e}")
                st.exception(e)
else:
    st.info("파일 두 개를 모두 업로드하면 처리 버튼이 활성화됩니다.")
